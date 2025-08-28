import openpyxl 
from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_file, send_from_directory, make_response
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
from io import BytesIO
import pandas as pd
import traceback
from werkzeug.utils import secure_filename
import shutil
from openpyxl.utils.dataframe import dataframe_to_rows
import subprocess
import sys
import json
from openpyxl.styles import Alignment
import numpy as np
from collections import Counter


app = Flask(__name__)
app.secret_key = os.urandom(24)  # Clave secreta segura

# Usuarios temporales (eliminar cuando se implemente BD)
USUARIOS_TEMPORALES = {
    'admin': {
        'password': 'admin1234',
        'roles': ['admin']  
    },
    'auxiliar': {
        'password': 'auxiliar1234',
        'roles': ['auxiliar']  
    },
    'arquitecto': {
        'password': 'arquitecto1234',
        'roles': ['arquitecto']  
    }
}
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
LOCAL_STORAGE_PATH = os.path.join(BASE_DIR, 'proyectos_data.json')  # Ruta del archivo JSON

# Configuración para Cupo Sombrilla
CUPO_SOMBRILLA_FOLDER = os.path.join(os.getcwd(), 'cupo_sombrilla_files')
if not os.path.exists(CUPO_SOMBRILLA_FOLDER):
    os.makedirs(CUPO_SOMBRILLA_FOLDER)

# Ruta para servir el favicon desde la raíz
@app.route('/LogoBancolombia.ico')
def favicon():
    return send_from_directory(os.path.join(app.root_path, 'static'),
                               'LogoBancolombia.ico', mimetype='image/vnd.microsoft.icon')

@app.route("/")
def index():
    return redirect(url_for('login'))

@app.route("/login", methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username').strip().lower()  # Normaliza el username
        password = request.form.get('password')
        role = request.form.get('role').strip().lower()  # Normaliza el rol
        
        # Validación de credenciales
        if username in USUARIOS_TEMPORALES:
            user_data = USUARIOS_TEMPORALES[username]
            
            if user_data['password'] == password:
                if role in user_data['roles']:
                    session['usuario'] = {
                        'nombre': username,
                        'rol': role
                    }
                    return jsonify({
                        'success': True,
                        'message': 'Autenticación exitosa',
                        'redirect': url_for('menu_principal')
                    })
                else:
                    return jsonify({
                        'success': False,
                        'message': f'Rol no válido. Roles permitidos: {", ".join(user_data["roles"])}'
                    }), 401
            else:
                return jsonify({
                    'success': False,
                    'message': 'Contraseña incorrecta'
                }), 401
        else:
            return jsonify({
                'success': False,
                'message': 'Usuario no encontrado'
            }), 401
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('usuario', None)
    return redirect(url_for('login'))

@app.route("/ruta_protegida")
def ruta_protegida():
    if 'usuario' not in session:
        return redirect(url_for('login'))
    
@app.route("/menu")
def menu_principal():
    if 'usuario' not in session:
        return redirect(url_for('login'))
    
    # Obtener datos del usuario desde la sesión
    usuario = session.get('usuario', {})
    
    return render_template("menuApp.html",
                         usuario_rol=usuario.get('rol', ''),
                         usuario_nombre=usuario.get('nombre', ''))  # Pasar variables correctamente

@app.route("/admin/roles")
def admin_roles():
    # Verificación más flexible de roles
    roles_permitidos = ['admin']  # Puedes agregar más roles
    
    if 'usuario' not in session:
        return redirect(url_for('login', next=request.url))
    
    if session.get('usuario', {}).get('rol') not in roles_permitidos:
        # Renderiza la misma plantilla pero con mensaje de error
        return render_template("pilotosTemplates/admin_roles.html",
                            es_admin=False,
                            acceso_denegado=True,
                            usuario=session['usuario'])
    
    return render_template("pilotosTemplates/admin_roles.html",
                         es_admin=True,
                         usuario=session['usuario'])

@app.route("/pilotos")
def modulo_pilotos():
    if 'usuario' not in session:
        return redirect(url_for('login'))
    # Opcional: Verificar si el rol tiene acceso
    if session['usuario']['rol'] not in ['admin', 'auxiliar', 'arquitecto']:
        return redirect(url_for('menu_principal'))
    return render_template("pilotosTemplates/pilotosApp.html")

@app.route("/ventacierta")
def venta_cierta():
    if 'usuario' not in session:
        return redirect(url_for('login'))
    # Opcional: Verificar si el rol tiene acceso
    if session['usuario']['rol'] not in ['admin', 'auxiliar', 'arquitecto']:
        return redirect(url_for('menu_principal'))
    return render_template("ventaCiertaTemplates/importar-datos.html")

@app.route("/crear_proyecto")
def crear_proyecto():
    if 'usuario' not in session:
        return redirect(url_for('login'))
    return render_template("pilotosTemplates/crear_proyecto.html")

@app.route("/config")
def config_app():
    if 'usuario' not in session:
        return redirect(url_for('login'))
        
    return render_template("pilotosTemplates/configApp.html")

@app.route("/seguimiento_proyecto")
def seguimiento_Proyecto():
    if 'usuario' not in session:
        return redirect(url_for('login'))
        
    return render_template("pilotosTemplates/seguimientoProyecto.html")

@app.route("/historicos_proyectos")
def historicos_Proyectos():
    if 'usuario' not in session:
        return redirect(url_for('login'))
        
    return render_template("pilotosTemplates/historico.html")

@app.route("/validacion_condiciones")
def validacion_condiciones():
    if 'usuario' not in session:
        return redirect(url_for('login'))
        
    return render_template("pilotosTemplates/validacion_condiciones.html")

# Configuración inicial - Cambiar esta ruta según necesidad
EXCEL_PATH = r'\\servidor\carpeta_compartida\proyectos_pilotos.xlsx'

def init_excel():
    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.append([
            'Fecha Creación', 'Nombre Proyecto', 'Titular', 'NIT/RUT', 'Ciudad',
            'Tipo Proyecto', 'Fecha Inicio', 'Monto Solicitado', 'Plazo (meses)',
            'Tasa Interés', 'Destino Recursos', 'Participantes', 'Documentos'
        ])
        wb.save(EXCEL_PATH)

@app.route('/guardar_proyecto', methods=['POST'])
def guardar_proyecto():
    try:
        data = request.json
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
        
        participantes = ", ".join([f"{p['nombre']} ({p['rol']} - {p['participacion']}%)" for p in data['participantes']])
        
        ws.append([
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            data['nombre_proyecto'],
            data['titular'],
            data['nit'],
            data['ciudad'],
            data['tipo_proyecto'],
            data['fecha_inicio'],
            data['monto_solicitado'],
            data['plazo'],
            data['tasa_interes'],
            data['destino_recursos'],
            participantes,
            str(data['documentos'])
        ])
        
        wb.save(EXCEL_PATH)
        return jsonify({'success': True, 'message': 'Proyecto guardado en Excel'})
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# Configuración inicial
EXCEL_PATH = r'\\servidor\carpeta_compartida\proyectos_pilotos.xlsx'

def generar_cupo_sombrilla(df_base_codigo, df_base_control, output_path):
    """Genera el archivo final de Cupo Sombrilla"""
    # Inicializar libro Excel
    wb = openpyxl.Workbook()
    
    # --- Hoja VISOR (Principal) ---
    ws_visor = wb.active
    ws_visor.title = "VISOR"
    
    # Añadir fórmulas clave
    ws_visor['D4'] = datetime.now().strftime('%Y-%m-%d')  # Fecha actualización
    ws_visor['D16'] = ""  # Cupo Sombilla Recomendado (se calculará después)
    ws_visor['H25'] = "=IFERROR(H20/(H20+H23), 0)"  # Wallet Share
    
    # --- Hoja PARAMETROS (Optimizada) ---
    ws_param = wb.create_sheet("PARAMETROS")
    # Ejemplo de estructura simplificada
    parametros_data = {
        'Variable': ['% Estrés', 'Participación Deuda Bancolombia', 'Promedio Financiación'],
        'Valor': [0.35, 0.60, 0.75]
    }
    df_param = pd.DataFrame(parametros_data)
    for r in dataframe_to_rows(df_param, index=False, header=True):
        ws_param.append(r)
    
    # --- Hoja base_codigo (Datos desde Python) ---
    ws_base_codigo = wb.create_sheet("base_codigo")
    for r in dataframe_to_rows(df_base_codigo, index=False, header=True):
        ws_base_codigo.append(r)
    
    # --- Hoja base_control (Controles) ---
    ws_base_control = wb.create_sheet("base_control")
    for r in dataframe_to_rows(df_base_control, index=False, header=True):
        ws_base_control.append(r)
    
    # --- Calcular y escribir Cupo Recomendado ---
    cupo_recomendado = calcular_cupo_recomendado(df_base_codigo, df_base_control)
    ws_visor['D16'] = cupo_recomendado
    
    # Guardar archivo final
    wb.save(output_path)

def calcular_cupo_recomendado(df_base_codigo, df_base_control):
    """Lógica compleja de cálculo (ejemplo simplificado)"""
    # Aquí implementarías la lógica real de negocio
    cupo_modelo = df_base_codigo.loc[0, 'Cupo_Modelo']
    ajustes = df_base_control.loc[0, 'Ajuste_GC']
    return cupo_modelo - ajustes

def aplicar_filtros(proyectos, filtros):
    filtered = proyectos
    
    if filtros.get('proyecto'):
        filtered = [p for p in filtered if filtros['proyecto'].lower() in p['nombre'].lower()]
    
    if filtros.get('titular'):
        filtered = [p for p in filtered if filtros['titular'].lower() in p['titular'].lower()]
    
    if filtros.get('nit'):
        filtered = [p for p in filtered if filtros['nit'] in p['nit']]
    
    if filtros.get('arquitecto'):
        filtered = [p for p in filtered if filtros['arquitecto'].lower() in p['arquitecto'].lower()]
    
    return filtered

def obtener_proyectos_desde_excel():
    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
        
        proyectos = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            arquitecto = next((p.split(' (')[0] for p in str(row[10]).split(', ') if 'Arquitecto' in p), '')
            
            proyecto = {
                'nombre': row[1] if len(row) > 1 else '',
                'titular': row[2] if len(row) > 2 else '',
                'nit': row[3] if len(row) > 3 else '',
                'arquitecto': arquitecto,
                'avance': 0,  # Campo temporal
                'ultimo_desembolso': row[0] if len(row) > 0 else ''
            }
            proyectos.append(proyecto)
        return proyectos
    except Exception as e:
        print(f"Error leyendo Excel: {str(e)}")
        return []

@app.route("/consulta_proyectos")
def consulta_proyectos():
    if 'usuario' not in session:
        return redirect(url_for('login'))
    
    proyectos = obtener_proyectos_desde_excel()
    filtros = request.args.to_dict()
    proyectos_filtrados = aplicar_filtros(proyectos, filtros)
    
    return render_template("pilotosTemplates/consulta_proyectos.html", proyectos=proyectos_filtrados)


# Configuración de rutas
EXPORTAR_RUTA = "C:\\CreditosConstructor"

# Asegurar que la ruta de exportación exista
if not os.path.exists(EXPORTAR_RUTA):
    os.makedirs(EXPORTAR_RUTA, exist_ok=True)

# ---------------- UTILIDADES ----------------

def consolidate_duplicate_columns(df):
    """
    Detecta columnas duplicadas en df y las consolida en una sola columna por nombre,
    tomando por fila el ultimo valor no nulo/no vacio entre las columnas duplicadas.
    Devuelve un DataFrame con nombres de columna unicos, manteniendo el orden original.
    """
    if df is None:
        return pd.DataFrame()
    if df.empty:
        return df.copy()

    cols = list(df.columns)
    counts = Counter(cols)
    dup_names = [name for name, cnt in counts.items() if cnt > 1]
    if not dup_names:
        return df.copy()

    consolidated = {}
    for name in dup_names:
        positions = [i for i, c in enumerate(cols) if c == name]
        sub = df.iloc[:, positions]

        def last_notnull(row):
            for v in reversed(row.values):
                if pd.notna(v):
                    if not (isinstance(v, str) and str(v).strip() == ''):
                        return v
            return np.nan

        serie = sub.apply(last_notnull, axis=1)
        consolidated[name] = serie

    new_df = pd.DataFrame(index=df.index)
    seen = set()
    for i, c in enumerate(cols):
        if c in dup_names:
            if c in seen:
                continue
            new_df[c] = consolidated[c].values
            seen.add(c)
        else:
            new_df[c] = df.iloc[:, i].values

    return new_df

def sanitize_dataframe_for_excel(df, date_format='%d/%m/%Y'):
    """
    Convierte columnas datetime / pandas.Timestamp a strings usando date_format.
    Reemplaza NaT/NaN por cadena vacia ''.
    Retorna copia.
    """
    if df is None:
        return pd.DataFrame()
    df2 = df.copy()
    for col in df2.columns:
        try:
            if pd.api.types.is_datetime64_any_dtype(df2[col]):
                df2[col] = df2[col].dt.strftime(date_format).where(df2[col].notna(), '')
            else:
                # convertir valores sueltos tipo Timestamp o datetime a string; NaN -> ''
                def conv_val(x):
                    if isinstance(x, (pd.Timestamp, datetime)):
                        try:
                            return x.strftime(date_format)
                        except Exception:
                            return ''
                    if pd.isna(x):
                        return ''
                    return x
                df2[col] = df2[col].apply(conv_val)
        except Exception:
            # si falla, dejar la columna tal cual
            pass
    return df2

def df_to_serializable_records(df, date_format='%Y-%m-%dT%H:%M:%S'):
    """
    Convierte df.to_dict(orient='records') pero transforma pandas.Timestamp/datetime a ISO-string
    y transforma NaN/NaT a None para un json.dump seguro.
    """
    if df is None:
        return []
    records = df.to_dict(orient='records')
    def conv(x):
        if isinstance(x, (pd.Timestamp, datetime)):
            try:
                return x.strftime(date_format)
            except Exception:
                return ""
        if pd.isna(x):
            return ""            # <-- devolver string vacío en vez de None
        if isinstance(x, str):
            return x.strip()
        return x
    safe = []
    for r in records:
        safe.append({k: conv(v) for k, v in r.items()})
    return safe

# ---------------- FUNCIONES HELPER PARA PERSISTENCIA ----------------
def leer_datos_completos():
    """Lee todos los datos desde el Excel principal"""
    excel_path = os.path.join(BASE_DIR, "data", "Base_principal_proyectos.xlsx")
    if not os.path.exists(excel_path):
        return {'proyectos': [], 'appData': {}}
    
    try:
        # Cargar la hoja "Proyectos"
        df = pd.read_excel(excel_path, sheet_name="Proyectos")

        # Reemplazar NaN con string vacío para evitar problemas en JSON
        df = df.fillna("")

        # Convertir a lista de diccionarios
        proyectos = df.to_dict(orient="records")

        # Normalizar estructura mínima
        for p in proyectos:
            if not isinstance(p, dict):
                continue
            p.setdefault('id_proyecto', p.get('id', ''))
            p.setdefault('participantes', [])
            if 'estado' in p and isinstance(p['estado'], str):
                p['estado'] = p['estado'].strip()
            else:
                p['estado'] = 'sin aprobar'

        return {'proyectos': proyectos, 'appData': {}}
    except Exception as e:
        print(f"Error leyendo Excel {excel_path}: {e}")
        return {'proyectos': [], 'appData': {}}

def guardar_datos_completos(datos):
    """Guarda todos los datos manteniendo estructura completa"""
    # Asegurar estructura correcta
    if not isinstance(datos, dict):
        datos = {'proyectos': datos, 'appData': {}}
    
    if 'proyectos' not in datos:
        datos['proyectos'] = []
    if 'appData' not in datos:
        datos['appData'] = {}
    
    # Escritura segura con archivo temporal
    tmp_path = LOCAL_STORAGE_PATH + '.tmp'
    try:
        with open(tmp_path, 'w', encoding='utf-8') as f:
            json.dump(datos, f, ensure_ascii=False, indent=2)
        
        # Reemplazar archivo original
        if os.path.exists(LOCAL_STORAGE_PATH):
            os.remove(LOCAL_STORAGE_PATH)
        os.rename(tmp_path, LOCAL_STORAGE_PATH)
        
        return True
    except Exception as e:
        print(f"Error guardando datos: {str(e)}")
        # Limpiar archivo temporal si existe
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        return False

# ---------------- FUNCION: subir_base_principal ----------------
@app.route('/subir_base_principal', methods=['POST'])
def subir_base_principal():
    try:
        # ===== Lista completa de columnas requeridas (sin cambios de nombres) =====
        COLUMNAS_COMPLETAS = [
            "id_proyecto", "fecha_creacion", "tipo_producto", "rango_proyecto", "grupoPrincipal",
            "subgrupo_1", "subgrupo_2", "subgrupo_3", "nit_grupo_riesgo", "nit_titular",
            "titular_credito", "nombre_proyecto", "tipo_inmuebles", "segmento", "ciudad",
            "tipo_fiducia", "fiduciaria", "gerente", "arquitecto", "auxiliar", "perito",
            "monto_solicitado_1_desembolso", "monto_solicitado_cpi", "monto_solicitado_lote",
            "total_valor_aprobado", "calificacion_it", "costos_financiables", "valor_lote",
            "valor_total_proyecto", "meses_programacion", "total_inmuebles", "meses_para_venta",
            "tipo_credito", "departamento", "caso_pcp_bizagi", "visitas_a_cobrar",
            "cobrar_estudio_tecnico", "fecha_aprobacion", "alerta_fecha_aprobacion",
            "vigencia_en_meses", "instancia_aprobacion", "condiciones_aprobacion",
            "porcentaje_solicitado_financiables", "validacion_licencia", "fecha_radicacion",
            "poliza_decenal", "caso_bizagi_juridico", "control_cruzado", "fecha_confirmacion",
            "fecha_primera_visita", "id_garantia", "meses_avanzados", "plazo_ajustado",
            "fecha_ultimo_informe_de_ventas", "fecha_inicio_de_ventas", "total_esperado_ventas",
            "inmuebles_vendidos_(unidades)", "valor_inmuebles_vendidos",
            "%_porcentaje_de_ventas_(unidades)", "%_porcentaje_de_ventas_(valor)",
            "valor_por_vender", "valor_canjes", "cartera_recaudada", "cartera_por_recaudar",
            "promedio_mensual", "promedio_trimestral", "validacion_fecha_infventas", "regional",
            "numero_visita", "fecha_visita", "valor_anticipos", "valor_anticipos_almacen",
            "programacion_inicial", "mes_inicio_obra", "avance_obra_meses",
            "avance_esperado_porcentaje", "avance_obra", "inversion_del_proyecto",
            "tipo_inmueble", "numero_de_inmuebles", "valor_garantia", "promedio_spi_trimestre",
            "spi_al_corte", "alerta_en_programación", "estado_ejecución_obra", "cpi_al_corte",
            "imprevistos_usados_a_la_fecha", "alerta_imprevistos_usados_a_la_fecha",
            "alerta_imprevistos_usados_vs_programación", "alerta_ejecutado_vs_invertido",
            "resumen_alerta_del_proyecto", "estado_ejecución_obra_mensual", "consumido_a_la_fecha",
            "imprevistos_y_reajustes_usados", "indirectos_invertidos_a_la_fecha",
            "indirectos_invertidos_en_%", "honorarios_invertidos_a_la_fecha",
            "honorarios_invertidos_en_%", "alerta_consumo_indirectos_vs_programacion",
            "alerta_consumo_honorarios_vs_programacion", "alerta_indirectos_vs_presupuesto",
            "alerta_honorarios_vs_presupuesto", "avance_obra_vs_avance_ci_honorarios_(cd)",
            "avance_obra_vs_avance_ci_honorarios_(ci+h)", "programacion_actual",
            "mes_terminacion_actual", "valla", "valor_entregado", "valor_por_entregar",
            "formula_a", "formula_b", "superavit", "cobertura", "maximo_desembolsar",
            "requiere_visto_bueno", "responsable_carga_desembolso", "fecha_carga_desembolso",
            "responsable_carga_venta", "responsable_carga_visita", "fecha_carga_venta",
            "fecha_carga_visita", "desembolsos_diarios", "valor_entregado_total"
            # Nuevas columnas
            "fecha_ultima_visita", "validacion_estado_ventas", "validacion_fecha_visita",
            "valor_desembolsar_preoperativo", "valor_desembolsar_constructor",
            "maximo_desembolsar_constructor", "maximo_desembolsar_superavit",
            "condiciones_especiales", "fecha_observacion_control", "valor_ampliacion",
            "observacion_control", 'condiciones_constructor', 'porcentaje_preoperativo','tipo_porcentaje_preoperativo',
            'porcentaje_constructor', 'tipo_porcentaje_constructor', 'porcentaje_recaudo_constructor', "estado"  
        ]

        MAPEO_COLUMNAS = {
            "ID PROYECTO": "id_proyecto",
            "TIPO CRED": "tipo_producto",
            "NUEVO TIPO DE PRODUCTO": "rango_proyecto",
            "NIT": "nit_titular",
            "TITULAR CREDITO": "titular_credito",
            "PROYECTO": "nombre_proyecto",
            "GERENTE": "gerente",
            "APROBADO 1° DESEMBOLSO": "monto_solicitado_1_desembolso",
            "APROBADO CPI": "monto_solicitado_cpi",
            "APROBADO LOTE": "monto_solicitado_lote",
            "VALOR APROBADO": "total_valor_aprobado",
            "COSTOS FINANCIABLES": "costos_financiables",
            "FECHA APROBACION (DD-MM-YY)": "fecha_aprobacion",
            "OBSERVACION (SE DEBE TENER EN CUENTA PARA PRIMER DESEMBOLSO)": "condiciones_aprobacion",
            "VISITAS A COBRAR": "visitas_a_cobrar",
            "COBRAR ESTUDIO TECNICO Y AVALUO": "cobrar_estudio_tecnico",
            "VALOR LOTE": "valor_lote",
            "VALOR GARANTIA": "valor_garantia",
            "VALOR TOTAL PROYECTO": "valor_total_proyecto",
            "FECHA VISITA AVANCE DE OBRA (DD-MM-YY)": "fecha_visita",
            "MESES PROGRAMACION DE OBRA": "meses_programacion",
            "MESES AVANZADOS DE OBRA": "meses_avanzados",
            "MESES PARA VENTA": "meses_para_venta",
            "PLAZO (ajustado según fecha visita)": "plazo_ajustado",
            "VALIDACION LICENCIA DE CONSTRUCCION": "validacion_licencia",
            "FECHA RADICACION LICENCIA ANTE CURADURIA": "fecha_radicacion",
            "REQUIERE POLIZA DECENAL (SI/NO)": "poliza_decenal",
            "FECHA ULTIMO INFORME DE VENTAS (DD-MM-YY)": "fecha_ultimo_informe_de_ventas",
            "TOTAL INMUEBLES": "total_inmuebles",
            "INMUEBLES VENDIDOS": "inmuebles_vendidos_(unidades)",
            "% PORCENTAJE DE VENTAS": "%_porcentaje_de_ventas_(unidades)",
            "VALIDACION ESTADO DE VENTAS": "validacion_estado_ventas",
            "FECHA ULTIMA VISITA AVANCE DE OBRA (DD-MM-YY)": "fecha_ultima_visita",
            "% ULTIMO AVANCE DE OBRA": "avance_obra",
            "VALOR INVERTIDO SEGÚN % AVANCE DE OBRA": "inversion_del_proyecto",
            "VALIDACION FECHA VISITA DE OBRA": "validacion_fecha_visita",
            "PERITO VISITAS": "perito",
            "VALOR ENTREGADO (Preoperativo y Constructor)": "valor_entregado",
            "VALOR X ENTREGAR": "valor_por_entregar",
            "DESEMBOLSO AUTORIZADO CON REGIMEN DE EXCEPCION (VR EN PESOS $)": "desembolsos_diarios",
            "VALOR A DESEMBOLSAR PREOPERATIVO": "valor_desembolsar_preoperativo",
            "VALOR A DESEMBOLSAR CREDITO LOTE": "monto_solicitado_lote",
            "VALOR MAXIMO A DESEMBOLSAR CONSTRUCTOR SEGUN % SOLICITADO": "maximo_desembolsar_constructor",
            "VALOR A DESEMBOLSAR CONSTRUCTOR SEGÚN % MAXIMO": "valor_desembolsar_constructor",
            "VALOR MÁXIMO A DESEMBOLSAR SEGÚN POLITICA DE SUPERÁVIT": "maximo_desembolsar_superavit",
            "VISTO BUENO": "requiere_visto_bueno",
            "RECOMENDACIÓN DESEMBOLSO SEGÚN POLÍTICA PARA DE SUPERÁVIT (COMENTARIO PARA AGREGAR EN BIZAGI) APLICA PARA PROYECTOS APROBADOS DESDE 23/05/2022": "superavit",
            "COBERTURA": "cobertura",
            "SUPERÁVIT": "superavit",
            "PCP BIZAGI": "caso_pcp_bizagi",
            "CONDICIONES ESPECIALES PARA DESEMBOLSAR, VALIDACION VISITAS, OTROS": "condiciones_especiales",
            "AQUITECTO GCC": "arquitecto",
            "AUXILIAR GCC": "auxiliar",
            "CIUDAD PROYECTO": "ciudad",
            "ID - NIT GRUPO DE RIESGO (GERENCIADOR)": "nit_grupo_riesgo",
            "GRUPO DE RIESGO - (GERENCIADOR CARPETA DE CLIENTES PYMES NACIONAL)": "grupoPrincipal",
            "ID GARANTIA (INICIAL Y/O DEFINITIVO)": "id_garantia",
            "INFORMACIÓN INICIAL Y/O DEFINITIVA HIPOTECA": "control_cruzado",
            "CASO BIZAGI JURÍDICO DE CONFIRMACIÓN": "caso_bizagi_juridico",
            "FECHA DE CONFIRMACIÓN ARCHIVO O BIZAGI": "fecha_confirmacion",
            "ESCRIBIR palabra LOTE PARA ESE TIPO DE CREDITO": "tipo_credito",
            "VALOR DESISTIDO": "valor_canjes",
            "VALOR AMPLIACION": "valor_ampliacion",
            "VALOR RECAUDADO": "cartera_recaudada",
            "VALOR POR RECAUDAR": "cartera_por_recaudar",
            "VALOR POR INVERTIR": "valor_por_entregar",
            "VALOR X VENDER": "valor_por_vender",
            "ANTICIPOS Y ALMACEN": "valor_anticipos_almacen",
            "ANÁLISIS INVERSIÓN": "inversion_del_proyecto",
            "FECHA INICIAL TERMINACIÓN": "fecha_inicio_de_ventas",
            "FECHA ACTUAL TERMINACIÓN": "fecha_ultimo_informe_de_ventas",
            "OBSERVACION CONTROL": "observacion_control",
            "FECHA OBSERVACION CONTROL": "fecha_observacion_control",
            "VIGENCIA APROBACIÓN (EN MESES)": "vigencia_en_meses",
            "ESTADO PROYECTO": "estado"
        }

        # ===== Validaciones iniciales =====
        if 'archivo' not in request.files:
            return jsonify({'success': False, 'message': 'No se encontro archivo'}), 400

        archivo = request.files['archivo']
        if archivo.filename == '':
            return jsonify({'success': False, 'message': 'Nombre de archivo invalido'}), 400

        # ===== Backup =====
        crear_backup = request.form.get('backup', 'true') == 'true'
        base_path = os.path.join(EXPORTAR_RUTA, 'Base_principal_proyectos.xlsx')

        if crear_backup and os.path.exists(base_path):
            fecha = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = os.path.join(EXPORTAR_RUTA, f'Backup_{fecha}.xlsx')
            shutil.copyfile(base_path, backup_path)

        # ===== LEER DATOS COMPLETOS EXISTENTES =====
        datos_completos = leer_datos_completos()
        proyectos_existentes = datos_completos.get('proyectos', [])
        app_data_existente = datos_completos.get('appData', {})

        df_local = pd.DataFrame(proyectos_existentes).fillna('') if proyectos_existentes else pd.DataFrame()
        if not df_local.empty and 'id_proyecto' in df_local.columns:
            df_local['id_proyecto'] = df_local['id_proyecto'].astype(str).str.strip()
            df_local = consolidate_duplicate_columns(df_local)

        # ===== Leer Excel subido (conservar columnas extras) =====
        try:
            xls = pd.ExcelFile(archivo)
            if 'Hoja1' not in xls.sheet_names:
                return jsonify({'success': False, 'message': 'El archivo no contiene una hoja llamada \"Hoja1\"'}), 400

            archivo.seek(0)
            df_nuevo = pd.read_excel(archivo, sheet_name='Hoja1', header=1, dtype={'ID PROYECTO': str})
            if 'ID PROYECTO' not in df_nuevo.columns:
                return jsonify({'success': False, 'message': 'El archivo no contiene la columna \"ID PROYECTO\"'}), 400

            df_nuevo = df_nuevo.rename(columns=MAPEO_COLUMNAS)
            df_nuevo = consolidate_duplicate_columns(df_nuevo)

            if 'id_proyecto' in df_nuevo.columns:
                df_nuevo['id_proyecto'] = df_nuevo['id_proyecto'].astype(str).str.strip()
                df_nuevo['id_proyecto'] = df_nuevo['id_proyecto'].replace('nan', '', regex=False)

            df_nuevo = df_nuevo[df_nuevo['id_proyecto'].astype(str).str.strip() != '']

            # formateo fechas conocidas si existen (string output)
            columnas_fecha = [
                'fecha_creacion', 'fecha_aprobacion', 'fecha_radicacion',
                'fecha_confirmacion', 'fecha_visita', 'fecha_ultimo_informe_de_ventas',
                'fecha_inicio_de_ventas', 'fecha_primera_visita', 'fecha_ultima_visita',
                'fecha_observacion_control'
            ]
            for col in columnas_fecha:
                if col in df_nuevo.columns:
                    valores_originales = df_nuevo[col].copy()
                    try:
                        tmp = pd.to_datetime(df_nuevo[col], errors='coerce', dayfirst=True)
                        mascara_valida = tmp.notna()
                        df_nuevo[col] = valores_originales.astype(str)
                        df_nuevo.loc[mascara_valida, col] = tmp.loc[mascara_valida].dt.strftime('%d/%m/%Y')
                    except Exception:
                        df_nuevo[col] = valores_originales

        except Exception as e:
            return jsonify({'success': False, 'message': f'Error procesando archivo: {str(e)}'}), 400
        

        # ===== Combinar: existente + local + nuevo =====
        if os.path.exists(base_path):
            try:
                df_existente_excel = pd.read_excel(base_path, sheet_name='Proyectos', dtype={'id_proyecto': str})
                df_existente_excel = df_existente_excel.fillna('')
                if 'id_proyecto' in df_existente_excel.columns:
                    df_existente_excel['id_proyecto'] = df_existente_excel['id_proyecto'].astype(str).str.strip()
                else:
                    df_existente_excel['id_proyecto'] = ''
                df_existente_excel = consolidate_duplicate_columns(df_existente_excel)

                frames = [df_existente_excel]
                if not df_local.empty:
                    frames.append(df_local)
                if not df_nuevo.empty:
                    frames.append(df_nuevo)
                df_concat = pd.concat(frames, ignore_index=True, sort=False).fillna('')
            except Exception as e:
                print(f"Error combinando datos: {str(e)}")
                if not df_local.empty:
                    df_concat = pd.concat([df_local, df_nuevo], ignore_index=True, sort=False).fillna('')
                else:
                    df_concat = df_nuevo.copy()
        else:
            if not df_local.empty:
                df_concat = pd.concat([df_local, df_nuevo], ignore_index=True, sort=False).fillna('')
            else:
                df_concat = df_nuevo.copy()

        # Consolidar duplicados que pudieron surgir en la concatenacion
        df_concat = consolidate_duplicate_columns(df_concat)
        df_concat = df_concat.astype(object)  # evita FutureWarning en replace

        # ===== Por id_proyecto tomar el ultimo valor NO VACIO por columna =====
        df_work = df_concat.replace(r'^\s*$', np.nan, regex=True)
        df_work = df_work.replace('nan', np.nan)

        if 'id_proyecto' in df_work.columns:
            def last_notnull(series):
                s = series.dropna()
                return s.iloc[-1] if not s.empty else np.nan

            group_key = df_work['id_proyecto'].astype(str).str.strip()
            df_grouped = df_work.groupby(group_key, dropna=False).agg(last_notnull)

            # previene el error de reset_index() si existe columna 'id_proyecto'
            if 'id_proyecto' in df_grouped.columns:
                df_grouped = df_grouped.drop(columns=['id_proyecto'])
            df_grouped.index.name = 'id_proyecto'
            df_combinado = df_grouped.reset_index()
        else:
            df_combinado = df_concat.copy()

        # asegurar columnas esperadas (sin eliminar extras)
        for col in COLUMNAS_COMPLETAS:
            if col not in df_combinado.columns:
                df_combinado[col] = ""


        cols_extra = [c for c in df_combinado.columns if c not in COLUMNAS_COMPLETAS]
        orden_final = COLUMNAS_COMPLETAS + cols_extra
        df_combinado = df_combinado[orden_final]

        # --- Normalizar observacion_control en el combinado ---
        if 'observacion_control' in df_combinado.columns:
            df_combinado['observacion_control'] = df_combinado['observacion_control'].astype(str).replace(['nan','None','NaT'], '', regex=True)
        else:
            df_combinado['observacion_control'] = ''


        # limpiar id_proyecto y filas vacias
        if 'id_proyecto' in df_combinado.columns:
            df_combinado['id_proyecto'] = df_combinado['id_proyecto'].astype(str).str.strip()
            df_combinado = df_combinado[df_combinado['id_proyecto'] != '']

        # ===== Guardar Excel (sanitizando fechas) =====
        try:
            df_for_excel = sanitize_dataframe_for_excel(df_combinado, date_format='%d/%m/%Y')
            with pd.ExcelWriter(base_path, engine='openpyxl') as writer:
                df_for_excel.to_excel(writer, index=False, sheet_name='Proyectos')
                worksheet = writer.sheets['Proyectos']
                try:
                    worksheet.protection.sheet = True
                    worksheet.protection.set_password('Riesgos2025*')
                    worksheet.protection.autoFilter = True
                    worksheet.protection.sort = True
                    worksheet.protection.insertRows = False
                    worksheet.protection.deleteRows = False
                except Exception as e:
                    print(f"No fue posible aplicar proteccion: {e}")

                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            val = '' if cell.value is None else str(cell.value)
                            if len(val) > max_length:
                                max_length = len(val)
                        except Exception:
                            pass
                    worksheet.column_dimensions[column_letter].width = min(max_length + 2, 150)
                try:
                    worksheet.freeze_panes = "A2"
                except Exception:
                    pass
        except Exception as e:
            return jsonify({'success': False, 'message': f'Error guardando archivo: {str(e)}'}), 500

        # ===== Actualizar LOCAL_STORAGE_PATH con estructura completa =====
        proyectos_guardar = []
        nuevos_cont = 0
        actualizados_cont = 0

        # --- Normalizar observacion_control ---
        if 'observacion_control' not in df_nuevo.columns:
            df_nuevo['observacion_control'] = ''
        else:
            df_nuevo['observacion_control'] = df_nuevo['observacion_control'].astype(str)
            df_nuevo['observacion_control'] = df_nuevo['observacion_control'].replace('nan', '', regex=False)

        try:
            df_combinado = consolidate_duplicate_columns(df_combinado)
            proyectos_guardar = df_to_serializable_records(df_combinado, date_format='%Y-%m-%dT%H:%M:%S')
            # asegurar id_proyecto strings
            for p in proyectos_guardar:
                p['id_proyecto'] = '' if p.get('id_proyecto') is None else str(p.get('id_proyecto')).strip()

            # Guardar estructura completa preservando appData
            contenido_completo = {
                'proyectos': proyectos_guardar,
                'appData': app_data_existente  # ← PRESERVA appData
            }
            
            guardar_datos_completos(contenido_completo)

            # conteos nuevos/actualizados segun df_existente y df_nuevo
            try:
                if 'df_existente_excel' in locals():
                    ids_existentes_previos_excel = set(str(x).strip() for x in df_existente_excel['id_proyecto'].dropna().astype(str))
                    ids_nuevos_excel = set(str(x).strip() for x in df_nuevo['id_proyecto'].dropna().astype(str)) if not df_nuevo.empty else set()
                    nuevos_cont = len(ids_nuevos_excel - ids_existentes_previos_excel)
                    actualizados_cont = len(ids_nuevos_excel & ids_existentes_previos_excel)
                else:
                    nuevos_cont = len(df_nuevo['id_proyecto'].dropna().astype(str).str.strip().unique()) if not df_nuevo.empty else 0
                    actualizados_cont = 0
            except Exception:
                nuevos_cont = 0
                actualizados_cont = 0

        except Exception as e:
            print(f"Error actualizando localStorage: {str(e)}")
            traceback.print_exc()

        return jsonify({
            'success': True,
            'message': 'Base principal actualizada (no pierde datos previos)',
            'nuevos': nuevos_cont,
            'actualizados': actualizados_cont,
            'proyectos': proyectos_guardar
        })
    except Exception as e:
        print(f"Error en subir_base_principal: {str(e)}\n{traceback.format_exc()}")
        return jsonify({'success': False, 'message': f'Error interno: {str(e)}'}), 500

# ---------------- FUNCION: exportar_excel ----------------
@app.route('/exportar_excel', methods=['POST'])
def exportar_excel():
    try:
        data = request.get_json(force=True)
        if not data:
            return jsonify({'success': False, 'message': 'Datos invalidos'}), 400

        proyectos = data.get('proyectos', [])
        if not proyectos:
            return jsonify({'success': False, 'message': 'No hay proyectos para exportar'}), 400

        # Normalizar payload y campos (mantener igual)
        campos_obligatorios = ['id_proyecto', 'nombre_proyecto', 'fecha_creacion']
        for proyecto in proyectos:
            if not all(key in proyecto for key in campos_obligatorios):
                return jsonify({'success': False, 'message': f'Falta campo obligatorio en proyecto {proyecto.get("id_proyecto", "")}'}), 400

            if 'condiciones_aprobacion' in proyecto:
                if isinstance(proyecto['condiciones_aprobacion'], str):
                    proyecto['condiciones_aprobacion'] = proyecto['condiciones_aprobacion'].split('; ')
                elif not isinstance(proyecto['condiciones_aprobacion'], list):
                    proyecto['condiciones_aprobacion'] = []

            if 'condiciones_constructor' in proyecto:
                if isinstance(proyecto['condiciones_constructor'], str):
                    proyecto['condiciones_constructor'] = proyecto['condiciones_constructor'].split('; ')
                elif not isinstance(proyecto['condiciones_constructor'], list):
                    proyecto['condiciones_constructor'] = []

            if 'observacion_control' in proyecto and proyecto['observacion_control']:
                proyecto['observacion_control'] = str(proyecto['observacion_control']).replace('\n', ' | ')
            else:
                proyecto['observacion_control'] = ''

            if 'desembolsos_diarios' in proyecto:
                if not isinstance(proyecto['desembolsos_diarios'], str):
                    proyecto['desembolsos_diarios'] = str(proyecto['desembolsos_diarios'])
            else:
                proyecto['desembolsos_diarios'] = 'Sin desembolsos diarios'

        df_nuevo = pd.DataFrame(proyectos).fillna('')

        # Procesar historiales (genericos, excepto comentarios)
        try:
            historiales = data.get('historiales', {})
            if isinstance(historiales, dict):
                for hist_key, hist_list in historiales.items():
                    if hist_key == 'historialComentarios':
                        continue
                    if not isinstance(hist_list, list) or not hist_list:
                        continue
                    dfh = pd.DataFrame(hist_list).fillna('')
                    if 'id_proyecto' not in dfh.columns:
                        continue

                    dfh['id_proyecto'] = dfh['id_proyecto'].astype(str).str.strip()

                    for pid, grupo in dfh.groupby('id_proyecto', dropna=False):
                        if pd.isna(pid) or str(pid).strip() == '':
                            continue
                        grupo_clean = grupo.replace(r'^\s*$', np.nan, regex=True)
                        if grupo_clean.dropna(axis=1, how='all').empty:
                            continue
                        ultima = grupo_clean.dropna(axis=1, how='all').iloc[-1]

                        mask = df_nuevo['id_proyecto'].astype(str).str.strip() == str(pid).strip()
                        if not mask.any():
                            continue
                        for col in ultima.index:
                            if col == 'id_proyecto':
                                continue
                            if col in df_nuevo.columns:
                                val = ultima[col]
                                df_nuevo.loc[mask, col] = val
        except Exception as e:
            print(f"Advertencia: no se pudo integrar historiales: {e}")

        # Leer datos existentes COMPLETOS
        datos_completos = leer_datos_completos()
        proyectos_existentes = datos_completos.get('proyectos', [])
        app_data_existente = datos_completos.get('appData', {})

        # Convertir proyectos existentes a DataFrame
        df_existente = pd.DataFrame(proyectos_existentes).fillna('') if proyectos_existentes else pd.DataFrame()
        if not df_existente.empty and 'id_proyecto' in df_existente.columns:
            df_existente['id_proyecto'] = df_existente['id_proyecto'].astype(str).str.strip()
            df_existente = consolidate_duplicate_columns(df_existente)

        # Convertir nuevos proyectos a DataFrame
        df_nuevo = pd.DataFrame(proyectos).fillna('')
        if not df_nuevo.empty and 'id_proyecto' in df_nuevo.columns:
            df_nuevo['id_proyecto'] = df_nuevo['id_proyecto'].astype(str).str.strip()
            df_nuevo = consolidate_duplicate_columns(df_nuevo)

        # COMBINAR DATOS EXISTENTES + NUEVOS
        if not df_existente.empty and not df_nuevo.empty:
            df_concat = pd.concat([df_existente, df_nuevo], ignore_index=True, sort=False).fillna('')
        elif not df_existente.empty:
            df_concat = df_existente.copy()
        elif not df_nuevo.empty:
            df_concat = df_nuevo.copy()
        else:
            df_concat = pd.DataFrame()

        # Consolidar duplicados
        df_concat = consolidate_duplicate_columns(df_concat)
        df_concat = df_concat.astype(object)
        df_work = df_concat.replace(r'^\s*$', np.nan, regex=True)
        df_work = df_work.replace('nan', np.nan)

        # Agrupar por id_proyecto y tomar último valor no nulo
        if 'id_proyecto' in df_work.columns and not df_work.empty:
            def last_notnull(s):
                s2 = s.dropna()
                return s2.iloc[-1] if not s2.empty else np.nan
            
            group_key = df_work['id_proyecto'].astype(str).str.strip()
            df_grouped = df_work.groupby(group_key, dropna=False).agg(last_notnull)
            
            if 'id_proyecto' in df_grouped.columns:
                df_grouped = df_grouped.drop(columns=['id_proyecto'])
            df_grouped.index.name = 'id_proyecto'
            df_combinado = df_grouped.reset_index()
        else:
            df_combinado = df_concat.copy()

        # Asegurar columnas mínimas
        columnas_esperadas = [
            'id_proyecto', 'nombre_proyecto', 'fecha_creacion',
            'condiciones_aprobacion', 'condiciones_constructor',
            'observacion_control', 'desembolsos_diarios'
        ]
        
        for col in columnas_esperadas:
            if col not in df_combinado.columns:
                df_combinado[col] = '' if col != 'observacion_control' else ''

        # Reordenar columnas
        cols_extra = [c for c in df_combinado.columns if c not in columnas_esperadas]
        orden_final = columnas_esperadas + cols_extra
        df_combinado = df_combinado[orden_final]

        # Limpiar filas vacías
        if 'id_proyecto' in df_combinado.columns:
            df_combinado['id_proyecto'] = df_combinado['id_proyecto'].astype(str).str.strip()
            df_combinado = df_combinado[df_combinado['id_proyecto'] != '']

        # Consolidar final
        df_combinado = consolidate_duplicate_columns(df_combinado)

        # 🔹 Reaplicar historialComentarios sobre df_combinado FINAL
        try:
            if isinstance(historiales, dict) and 'historialComentarios' in historiales:
                comentarios = historiales['historialComentarios']
                if isinstance(comentarios, list):
                    for pid in set(str(c.get('idProyecto')).strip() for c in comentarios if c.get('idProyecto')):
                        comentariosProyecto = [
                            c for c in comentarios if str(c.get('idProyecto')).strip() == pid
                        ]
                        comentariosProyecto.sort(key=lambda x: x.get('fecha') or '')

                        def fmt_fecha(fecha_raw):
                            try:
                                fecha = pd.to_datetime(fecha_raw)
                                return fecha.strftime("%d/%m/%Y %H:%M")
                            except Exception:
                                return str(fecha_raw)

                        texto = "\n".join(
                            f"{c.get('usuario','')}: {c.get('comentario','')} ({fmt_fecha(c.get('fecha'))})"
                            for c in comentariosProyecto
                        )
                        mask = df_combinado['id_proyecto'].astype(str).str.strip() == pid
                        if mask.any():
                            df_combinado.loc[mask, 'observacion_control'] = texto
        except Exception as e:
            print(f"Error integrando historialComentarios al Excel: {e}")

        # Guardar Excel
        try:
            df_for_excel = sanitize_dataframe_for_excel(df_combinado, date_format='%d/%m/%Y')
            base_path = os.path.join(EXPORTAR_RUTA, 'Base_principal_proyectos.xlsx')
            
            with pd.ExcelWriter(base_path, engine='openpyxl') as writer:
                df_for_excel.to_excel(writer, index=False, sheet_name='Proyectos')
                worksheet = writer.sheets['Proyectos']
                
                try:
                    worksheet.protection.sheet = True
                    worksheet.protection.set_password('Riesgos2025*')
                    worksheet.protection.autoFilter = True
                    worksheet.protection.sort = True
                    worksheet.protection.insertRows = False
                    worksheet.protection.deleteRows = False
                except Exception as e:
                    print(f"No fue posible aplicar proteccion: {e}")

                for column in worksheet.columns:
                    try:
                        max_length = max(
                            len(str(cell.value)) if cell.value else 0
                            for cell in column
                        )
                        worksheet.column_dimensions[column[0].column_letter].width = min(max_length + 2, 150)
                    except Exception:
                        pass
                
                try:
                    worksheet.freeze_panes = "A2"
                except Exception:
                    pass
        except Exception as e:
            return jsonify({'success': False, 'message': f'Error guardando Excel: {str(e)}'}), 500

        # Actualizar LOCAL_STORAGE_PATH con estructura COMPLETA
        proyectos_guardar = []
        nuevos_cont = 0
        actualizados_cont = 0
        
        try:
            df_combinado = consolidate_duplicate_columns(df_combinado)
            proyectos_guardar = df_to_serializable_records(df_combinado, date_format='%Y-%m-%dT%H:%M:%S')
            
            for p in proyectos_guardar:
                p['id_proyecto'] = '' if p.get('id_proyecto') is None else str(p.get('id_proyecto')).strip()
                p['estado'] = p.get('estado', 'sin aprobar')

            contenido_completo = {
                'proyectos': proyectos_guardar,
                'appData': app_data_existente
            }
            guardar_datos_completos(contenido_completo)

            ids_existentes_previos = set(str(p.get('id_proyecto', '')).strip() for p in proyectos_existentes)
            ids_nuevos = set(str(p.get('id_proyecto', '')).strip() for p in proyectos)
            
            nuevos_cont = len(ids_nuevos - ids_existentes_previos)
            actualizados_cont = len(ids_nuevos & ids_existentes_previos)

        except Exception as e:
            print(f"Error actualizando localStorage: {str(e)}")
            traceback.print_exc()

        # Conteos para Excel
        try:
            if not df_existente.empty:
                ids_existentes_previos_excel = set(str(x).strip() for x in df_existente['id_proyecto'].dropna().astype(str))
                ids_nuevos_excel = set(str(x).strip() for x in df_nuevo['id_proyecto'].dropna().astype(str)) if not df_nuevo.empty else set()
                nuevos_excel = len(ids_nuevos_excel - ids_existentes_previos_excel)
                actualizados_excel = len(ids_nuevos_excel & ids_existentes_previos_excel)
            else:
                nuevos_excel = len(df_nuevo['id_proyecto'].dropna().astype(str).str.strip().unique()) if not df_nuevo.empty else 0
                actualizados_excel = 0
        except Exception:
            nuevos_excel = 0
            actualizados_excel = 0

        response = {
            'success': True,
            'message': 'Base principal exportada y actualizada (preserva valores previos no vacios)',
            'nuevos': nuevos_cont,
            'actualizados': actualizados_cont,
            'nuevos_excel': nuevos_excel,
            'actualizados_excel': actualizados_excel,
            'archivo_base': os.path.basename(base_path),
            'proyectos': proyectos_guardar
        }
        return jsonify(response), 200

    except Exception as e:
        print(f"Error en exportar_excel: {str(e)}\n{traceback.format_exc()}")
        return jsonify({'success': False, 'message': f'Error interno: {str(e)}'}), 500



@app.route('/importar_excel', methods=['POST'])
def importar_excel():
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No se subió archivo'}), 400
            
        file = request.files['file']
        if file.filename.endswith('.xlsx'):
            file.save(EXCEL_PATH)
            return jsonify({'success': True, 'message': 'Datos actualizados desde Excel'})
            
        return jsonify({'success': False, 'error': 'Formato no válido'}), 400
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    
@app.route("/cupo_sombrilla")
def modulo_cupo_sombrilla():
    if 'usuario' not in session:
        return redirect(url_for('login'))
    return render_template("cupoSombrillaTemplates/cupoSombrilla.html")


CUPO_SOMBRILLA_FOLDER = r"C:\CreditosConstructor\cupo_sombrilla_files"

@app.route('/cupo_sombrilla/run_script', methods=['POST'])
def run_cupo_sombrilla_script():
    try:
        # Ruta directa según tu estructura real
        script_path = r"C:\CreditosConstructor\Script_Cupo_Sombrilla_V1.py"
        
        # Verificar existencia del script
        if not os.path.exists(script_path):
            return jsonify({
                'success': False,
                'error': f'Script no encontrado en: {script_path}'
            }), 404

        # Crear carpeta de destino si no existe
        os.makedirs(CUPO_SOMBRILLA_FOLDER, exist_ok=True)

        # Ejecutar script
        python_exec = sys.executable
        app.logger.info(f"Ejecutando script: {script_path}")
        
        result = subprocess.run(
            [python_exec, script_path],
            cwd=os.path.dirname(script_path),
            capture_output=True,
            text=True,
            timeout=300
        )

        # Verificar resultados
        if result.returncode != 0:
            error_details = {
                "returncode": result.returncode,
                "stdout": result.stdout,
                "stderr": result.stderr
            }
            return jsonify({
                'success': False,
                'error': 'Error en ejecución del script',
                'details': error_details
            }), 500

        # Buscar archivos generados
        generated_files = []
        expected_files = [
            'BASE_Sombrilla.xlsx',
            'creditlean.xlsx',
            'cenegar_Saldos.xlsx'
        ]
        
        for filename in expected_files:
            src = os.path.join(os.path.dirname(script_path), filename)
            if os.path.exists(src):
                dest = os.path.join(CUPO_SOMBRILLA_FOLDER, filename)
                shutil.move(src, dest)
                generated_files.append(filename)
            else:
                app.logger.warning(f"Archivo esperado no generado: {filename}")

        if not generated_files:
            return jsonify({
                'success': False,
                'error': 'No se generaron archivos de salida'
            }), 500

        return jsonify({
            'success': True,
            'message': 'Script ejecutado correctamente',
            'files': generated_files
        })

    except subprocess.TimeoutExpired:
        return jsonify({
            'success': False,
            'error': 'Tiempo de ejecución excedido (5 minutos)'
        }), 500
        
    except Exception as e:
        error_trace = traceback.format_exc()
        app.logger.error(f"Error inesperado: {str(e)}\n{error_trace}")
        return jsonify({
            'success': False,
            'error': f'Error interno: {str(e)}',
            'trace': error_trace
        }), 500
    
@app.route('/cupo_sombrilla/process_step2', methods=['POST'])
def process_step2():
    try:
        cenegar_file = request.files['cenegar_file']
        if not cenegar_file:
            return jsonify({'success': False, 'error': 'No se subió archivo cenegar_saldos'}), 400
        
        # Guardar archivo
        cenegar_path = os.path.join(CUPO_SOMBRILLA_FOLDER, 'cenegar_saldos_processed.xlsx')
        cenegar_file.save(cenegar_path)
        
        # Procesamiento adicional (ejemplo)
        df = pd.read_excel(cenegar_path)
        
        # 1. Convertir columna C a numérico
        df['Columna_C'] = pd.to_numeric(df['Columna_C'], errors='coerce')
        
        # 2. Insertar columna para identificar crédito constructor
        df.insert(loc=3, column='Tipo_Credito', value='')
        
        # 3. Cruzar con archivo de saldos (simulado)
        # En producción esto sería una operación real con otro archivo
        df['Es_Constructor'] = df.apply(lambda row: 'ND' 
                                        if row['Columna_C'] > 1000000 
                                        else 'Constructor', axis=1)
        
        # 4. Filtrar solo ND (deuda corporativa)
        df = df[df['Es_Constructor'] == 'ND']
        
        # Guardar resultado
        df.to_excel(cenegar_path, index=False)
        
        return jsonify({'success': True, 'message': 'Paso 2 procesado correctamente'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# Ruta para procesar el Paso 3 (ahora incluye la generación completa)
@app.route('/cupo_sombrilla/process_step3', methods=['POST'])
def process_step3():
    try:
        # Obtener archivos subidos en pasos anteriores
        base_path = os.path.join(CUPO_SOMBRILLA_FOLDER, 'BASE_Sombrilla_processed.xlsx')
        control_path = os.path.join(CUPO_SOMBRILLA_FOLDER, 'cenegar_saldos_processed.xlsx')
        
        if not os.path.exists(base_path) or not os.path.exists(control_path):
            return jsonify({
                'success': False, 
                'error': 'Faltan archivos procesados de pasos anteriores'
            }), 400
        
        # Cargar datos procesados
        df_base_codigo = pd.read_excel(base_path)
        df_base_control = pd.read_excel(control_path)
        
        # Generar archivo final
        result_path = os.path.join(CUPO_SOMBRILLA_FOLDER, 'CUPO_SOMBRILLA_AUTOMATIZADO_V2.xlsx')
        generar_cupo_sombrilla(df_base_codigo, df_base_control, result_path)
        
        return jsonify({
            'success': True, 
            'message': 'Archivo de Cupo Sombrilla generado correctamente',
            'file': 'CUPO_SOMBRILLA_AUTOMATIZADO_V2.xlsx'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/cupo_sombrilla/download/<filename>', methods=['GET'])
def download_cupo_file(filename):
    try:
        return send_from_directory(
            directory=CUPO_SOMBRILLA_FOLDER,
            path=filename,
            as_attachment=True
        )
    except Exception as e:
        return str(e), 404

@app.route('/api/proyectos', methods=['GET'])
def obtener_proyectos_api():
    try:
        datos_completos = leer_datos_completos()
        proyectos = datos_completos.get('proyectos', [])

        # Normalizar campos mínimos por proyecto
        for p in proyectos:
            if not isinstance(p, dict):
                continue
            p.setdefault('id_proyecto', p.get('id', ''))
            p.setdefault('participantes', [])
            # Normalizar estado: respetar si ya existe
            if 'estado' in p and isinstance(p['estado'], str):
                p['estado'] = p['estado'].strip()
            else:
                p['estado'] = 'sin aprobar'
        return jsonify(proyectos)
    except Exception as e:
        app.logger.exception("Error leyendo proyectos")
        return jsonify([])  # fallback: lista vacia

# Reemplaza la funcion actual por esta en appRun.py
@app.route('/api/proyectos/<proyecto_id>', methods=['GET'])
def obtener_proyecto_por_id(proyecto_id):
    try:
        datos_completos = leer_datos_completos()
        proyectos = datos_completos.get('proyectos', [])
        
        # Normalizar búsqueda considerando ambos campos posibles
        target_id = str(proyecto_id).strip().lower()
        
        for proyecto in proyectos:
            # Verificar ambos campos posibles para el ID
            id_candidates = [
                str(proyecto.get('id_proyecto', '')).strip().lower(),
                str(proyecto.get('id', '')).strip().lower()
            ]
            
            if target_id in id_candidates:
                # Asegurar campos necesarios
                proyecto.setdefault('participantes', [])
                proyecto.setdefault('gestiones', [])
                
                if 'estado' not in proyecto:
                    proyecto['estado'] = 'sin aprobar'
                elif isinstance(proyecto['estado'], str):
                    proyecto['estado'] = proyecto['estado'].strip()
                
                return jsonify({
                    'success': True, 
                    'proyecto': proyecto
                })
        
        return jsonify({
            'success': False, 
            'message': 'Proyecto no encontrado'
        }), 404
        
    except Exception as e:
        app.logger.error(f"Error al obtener proyecto: {str(e)}")
        return jsonify({
            'success': False, 
            'message': f'Error interno: {str(e)}'
        }), 500    

# Cambiar todas las instancias donde se busca por 'id' a 'id_proyecto'

@app.route('/api/proyectos/<proyecto_id>', methods=['DELETE'])
def api_eliminar_proyecto(proyecto_id):
    try:
        datos_completos = leer_datos_completos()
        proyectos = datos_completos.get('proyectos', [])
        app_data = datos_completos.get('appData', {})
        
        # Buscar por id_proyecto en lugar de id
        proyectos_actualizados = [p for p in proyectos if p.get('id_proyecto') != proyecto_id]
        
        if len(proyectos_actualizados) == len(proyectos):
            return jsonify({'success': False, 'message': 'Proyecto no encontrado'}), 404
        
        # Guardar estructura completa
        guardar_datos_completos({
            'proyectos': proyectos_actualizados,
            'appData': app_data
        })
        
        return jsonify({'success': True, 'message': 'Proyecto eliminado correctamente'})
    
    except Exception as e:
        print(f"Error al eliminar proyecto: {str(e)}")
        return jsonify({'success': False, 'message': f'Error interno: {str(e)}'}), 500


# Aplicar el mismo cambio a los otros endpoints relacionados con gestiones
@app.route('/api/proyectos/<proyecto_id>/gestiones', methods=['POST'])
def guardar_gestion_proyecto(proyecto_id):
    try:
        datos_gestion = request.get_json()
        
        datos_completos = leer_datos_completos()
        proyectos = datos_completos.get('proyectos', [])
        app_data = datos_completos.get('appData', {})
        
        # Buscar proyecto por ID
        proyecto_index = next((i for i, p in enumerate(proyectos) if str(p.get('id_proyecto', '')).strip() == proyecto_id.strip()), None)
        
        if proyecto_index is None:
            return jsonify({'success': False, 'message': 'Proyecto no encontrado'}), 404
        
        # Inicializar array de gestiones si no existe
        if 'gestiones' not in proyectos[proyecto_index]:
            proyectos[proyecto_index]['gestiones'] = []
        
        # Agregar fecha y usuario a la gestión
        datos_gestion['fecha'] = datetime.now().isoformat()
        datos_gestion['usuario'] = session.get('usuario', {}).get('nombre', 'Usuario desconocido')
        
        # Guardar gestión
        proyectos[proyecto_index]['gestiones'].append(datos_gestion)
        
        # Si es una gestión de vigencia, actualizar el campo correspondiente
        if datos_gestion.get('tipo') == 'vigencia' and 'nuevo' in datos_gestion:
            proyectos[proyecto_index]['vigencia_en_meses'] = datos_gestion['nuevo']
        
        # Guardar cambios manteniendo appData
        guardar_datos_completos({
            'proyectos': proyectos,
            'appData': app_data
        })
        
        return jsonify({'success': True, 'message': 'Gestión guardada correctamente'})
    
    except Exception as e:
        print(f"Error al guardar gestión: {str(e)}")
        return jsonify({'success': False, 'message': f'Error interno: {str(e)}'}), 500

@app.route('/api/proyectos/<proyecto_id>/gestiones/<gestion_id>', methods=['DELETE'])
def api_eliminar_gestion(proyecto_id, gestion_id):
    try:
        datos_completos = leer_datos_completos()
        proyectos = datos_completos.get('proyectos', [])
        app_data = datos_completos.get('appData', {})
        
        # Buscar por id_proyecto
        proyecto_index = next((i for i, p in enumerate(proyectos) if p.get('id_proyecto') == proyecto_id), None)
        
        if proyecto_index is None:
            return jsonify({'success': False, 'message': 'Proyecto no encontrado'}), 404
        
        if 'gestiones' in proyectos[proyecto_index]:
            gestiones_actualizadas = [g for g in proyectos[proyecto_index]['gestiones'] if g.get('id') != gestion_id]
            
            if len(gestiones_actualizadas) == len(proyectos[proyecto_index]['gestiones']):
                return jsonify({'success': False, 'message': 'Gestión no encontrada'}), 404
            
            proyectos[proyecto_index]['gestiones'] = gestiones_actualizadas
        
        # Guardar manteniendo appData
        guardar_datos_completos({
            'proyectos': proyectos,
            'appData': app_data
        })
        
        return jsonify({'success': True, 'message': 'Gestión eliminada correctamente'})
    
    except Exception as e:
        print(f"Error al eliminar gestión: {str(e)}")
        return jsonify({'success': False, 'message': f'Error interno: {str(e)}'}), 500

# Endpoint para obtener datos de la aplicación
@app.route('/api/appData', methods=['GET'])
def obtener_app_data():
    datos_completos = leer_datos_completos()
    app_data = datos_completos.get('appData', {})
    
    # Asegurar estructura mínima
    default_appdata = {
        'seguimientos': [],
        'historialVentas': [],
        'historialVisitas': [],
        'historialDesembolsos': [],
        'historialDesembolsosDiarios': [],
        'historialComentarios': []
    }
    
    # Combinar con defaults si faltan campos
    for key in default_appdata:
        if key not in app_data:
            app_data[key] = default_appdata[key]
    
    return jsonify(app_data)


@app.route('/api/appData', methods=['PUT'])
def guardar_app_data():
    try:
        nuevos_datos = request.get_json()
        if nuevos_datos is None:
            return jsonify({'success': False, 'message': 'Payload vacio'}), 400

        datos_completos = leer_datos_completos()
        proyectos = datos_completos.get('proyectos', [])
        
        # Actualizar solo appData, preservando proyectos
        guardar_datos_completos({
            'proyectos': proyectos,
            'appData': nuevos_datos
        })

        return jsonify({'success': True, 'appData': nuevos_datos})
    except Exception as e:
        app.logger.exception("Error guardando appData")
        return jsonify({'success': False, 'error': str(e)}), 500

    
@app.route('/api/proyectos', methods=['PUT'])
def guardar_proyectos_api():
    try:
        proyectos_nuevos = request.get_json()
        if proyectos_nuevos is None:
            return jsonify({'success': False, 'message': 'Payload vacio'}), 400

        datos_completos = leer_datos_completos()
        app_data = datos_completos.get('appData', {})
        
        # Actualizar solo proyectos, preservando appData
        guardar_datos_completos({
            'proyectos': proyectos_nuevos,
            'appData': app_data
        })

        return jsonify({'success': True, 'proyectos': proyectos_nuevos})
    except Exception as e:
        app.logger.exception("Error guardando proyectos")
        return jsonify({'success': False, 'error': str(e)}), 500
    
@app.route('/api/proyectos/<proyecto_id>/estado', methods=['PUT'])
def actualizar_estado_proyecto(proyecto_id):
    try:
        payload = request.get_json(force=True)
        nuevo_estado = payload.get('estado') if isinstance(payload, dict) else None
        if not nuevo_estado:
            return jsonify({'success': False, 'message': 'Falta campo estado'}), 400

        datos_completos = leer_datos_completos()
        proyectos = datos_completos.get('proyectos', [])
        app_data = datos_completos.get('appData', {})

        # Buscar proyecto y actualizar estado
        encontrado = False
        for p in proyectos:
            if str(p.get('id_proyecto', '') or p.get('id', '')).strip() == str(proyecto_id).strip():
                p['estado'] = nuevo_estado
                encontrado = True
                break

        if not encontrado:
            return jsonify({'success': False, 'message': 'Proyecto no encontrado'}), 404

        # Guardar estructura completa
        guardar_datos_completos({
            'proyectos': proyectos,
            'appData': app_data
        })

        return jsonify({'success': True, 'proyecto_id': proyecto_id, 'nuevo_estado': nuevo_estado})
    except Exception as e:
        app.logger.exception("Error actualizando estado proyecto")
        return jsonify({'success': False, 'error': str(e)}), 500
      
    
if __name__ == "__main__":
     app.run(host="0.0.0.0", port=5000)